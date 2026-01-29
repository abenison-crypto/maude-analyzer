"""Events API router."""

from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import date
import io
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from api.services.queries import QueryService
from api.services.filters import DeviceFilters
from api.models.schemas import (
    EventListResponse,
    EventDetail,
    StatsResponse,
    ManufacturerItem,
    ProductCodeItem,
)

router = APIRouter()


def parse_device_filters(
    brand_names: Optional[str] = None,
    generic_names: Optional[str] = None,
    device_manufacturers: Optional[str] = None,
    model_numbers: Optional[str] = None,
    implant_flag: Optional[str] = None,
    device_product_codes: Optional[str] = None,
) -> Optional[DeviceFilters]:
    """Parse device filter query parameters into DeviceFilters object."""
    has_filters = any([
        brand_names, generic_names, device_manufacturers,
        model_numbers, implant_flag, device_product_codes
    ])

    if not has_filters:
        return None

    return DeviceFilters(
        brand_names=brand_names.split(",") if brand_names else None,
        generic_names=generic_names.split(",") if generic_names else None,
        device_manufacturers=device_manufacturers.split(",") if device_manufacturers else None,
        model_numbers=model_numbers.split(",") if model_numbers else None,
        implant_flag=implant_flag if implant_flag in ('Y', 'N') else None,
        device_product_codes=device_product_codes.split(",") if device_product_codes else None,
    )


@router.get("", response_model=EventListResponse)
async def list_events(
    # Core filters
    manufacturers: Optional[str] = Query(None, description="Comma-separated manufacturer names"),
    product_codes: Optional[str] = Query(None, description="Comma-separated product codes"),
    event_types: Optional[str] = Query(None, description="Comma-separated event types (D,I,M,O)"),
    date_from: Optional[date] = Query(None, description="Start date"),
    date_to: Optional[date] = Query(None, description="End date"),
    search_text: Optional[str] = Query(None, description="Search text"),
    # Device filters
    brand_names: Optional[str] = Query(None, description="Comma-separated device brand names"),
    generic_names: Optional[str] = Query(None, description="Comma-separated device generic names"),
    device_manufacturers: Optional[str] = Query(None, description="Comma-separated device manufacturer names"),
    model_numbers: Optional[str] = Query(None, description="Comma-separated device model numbers"),
    implant_flag: Optional[str] = Query(None, description="Implant flag (Y/N)"),
    device_product_codes: Optional[str] = Query(None, description="Comma-separated device product codes"),
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=1000, description="Results per page"),
):
    """List MDR events with filtering and pagination."""
    query_service = QueryService()

    # Parse comma-separated values
    mfr_list = manufacturers.split(",") if manufacturers else None
    code_list = product_codes.split(",") if product_codes else None
    type_list = event_types.split(",") if event_types else None

    # Parse device filters
    device_filters = parse_device_filters(
        brand_names=brand_names,
        generic_names=generic_names,
        device_manufacturers=device_manufacturers,
        model_numbers=model_numbers,
        implant_flag=implant_flag,
        device_product_codes=device_product_codes,
    )

    result = query_service.get_events(
        manufacturers=mfr_list,
        product_codes=code_list,
        event_types=type_list,
        date_from=date_from,
        date_to=date_to,
        search_text=search_text,
        device_filters=device_filters,
        page=page,
        page_size=page_size,
    )

    return result


@router.get("/stats", response_model=StatsResponse)
async def get_stats(
    # Core filters
    manufacturers: Optional[str] = Query(None, description="Comma-separated manufacturer names"),
    product_codes: Optional[str] = Query(None, description="Comma-separated product codes"),
    event_types: Optional[str] = Query(None, description="Comma-separated event types (D,I,M,O)"),
    date_from: Optional[date] = Query(None, description="Start date"),
    date_to: Optional[date] = Query(None, description="End date"),
    # Device filters
    brand_names: Optional[str] = Query(None, description="Comma-separated device brand names"),
    generic_names: Optional[str] = Query(None, description="Comma-separated device generic names"),
    device_manufacturers: Optional[str] = Query(None, description="Comma-separated device manufacturer names"),
    model_numbers: Optional[str] = Query(None, description="Comma-separated device model numbers"),
    implant_flag: Optional[str] = Query(None, description="Implant flag (Y/N)"),
    device_product_codes: Optional[str] = Query(None, description="Comma-separated device product codes"),
):
    """Get summary statistics for events matching filters."""
    query_service = QueryService()

    mfr_list = manufacturers.split(",") if manufacturers else None
    code_list = product_codes.split(",") if product_codes else None
    type_list = event_types.split(",") if event_types else None

    # Parse device filters
    device_filters = parse_device_filters(
        brand_names=brand_names,
        generic_names=generic_names,
        device_manufacturers=device_manufacturers,
        model_numbers=model_numbers,
        implant_flag=implant_flag,
        device_product_codes=device_product_codes,
    )

    return query_service.get_event_stats(
        manufacturers=mfr_list,
        product_codes=code_list,
        event_types=type_list,
        date_from=date_from,
        date_to=date_to,
        device_filters=device_filters,
    )


@router.get("/manufacturers", response_model=list[ManufacturerItem])
async def list_manufacturers(
    search: Optional[str] = Query(None, description="Search term"),
    limit: int = Query(100, ge=1, le=500, description="Max results"),
):
    """Get list of manufacturers for autocomplete."""
    query_service = QueryService()
    return query_service.get_manufacturer_list(search=search, limit=limit)


@router.get("/product-codes", response_model=list[ProductCodeItem])
async def list_product_codes(
    search: Optional[str] = Query(None, description="Search term"),
    limit: int = Query(100, ge=1, le=500, description="Max results"),
):
    """Get list of product codes for autocomplete."""
    query_service = QueryService()
    return query_service.get_product_code_list(search=search, limit=limit)


@router.get("/export")
async def export_events(
    # Core filters
    manufacturers: Optional[str] = Query(None, description="Comma-separated manufacturer names"),
    product_codes: Optional[str] = Query(None, description="Comma-separated product codes"),
    event_types: Optional[str] = Query(None, description="Comma-separated event types (D,I,M,O)"),
    date_from: Optional[date] = Query(None, description="Start date"),
    date_to: Optional[date] = Query(None, description="End date"),
    search_text: Optional[str] = Query(None, description="Search text"),
    # Device filters
    brand_names: Optional[str] = Query(None, description="Comma-separated device brand names"),
    generic_names: Optional[str] = Query(None, description="Comma-separated device generic names"),
    device_manufacturers: Optional[str] = Query(None, description="Comma-separated device manufacturer names"),
    model_numbers: Optional[str] = Query(None, description="Comma-separated device model numbers"),
    implant_flag: Optional[str] = Query(None, description="Implant flag (Y/N)"),
    device_product_codes: Optional[str] = Query(None, description="Comma-separated device product codes"),
    # Export options
    format: str = Query("csv", description="Export format (csv or xlsx)"),
    max_records: int = Query(10000, ge=1, le=100000, description="Maximum records to export"),
):
    """Export events to CSV or Excel format."""
    query_service = QueryService()

    mfr_list = manufacturers.split(",") if manufacturers else None
    code_list = product_codes.split(",") if product_codes else None
    type_list = event_types.split(",") if event_types else None

    # Parse device filters
    device_filters = parse_device_filters(
        brand_names=brand_names,
        generic_names=generic_names,
        device_manufacturers=device_manufacturers,
        model_numbers=model_numbers,
        implant_flag=implant_flag,
        device_product_codes=device_product_codes,
    )

    # Get events (limited to max_records)
    result = query_service.get_events(
        manufacturers=mfr_list,
        product_codes=code_list,
        event_types=type_list,
        date_from=date_from,
        date_to=date_to,
        search_text=search_text,
        device_filters=device_filters,
        page=1,
        page_size=max_records,
    )

    events = result["events"]
    headers = [
        "MDR Report Key", "Report Number", "Date Received", "Date of Event",
        "Event Type", "Manufacturer", "Product Code"
    ]

    if format == "xlsx":
        if not EXCEL_AVAILABLE:
            raise HTTPException(status_code=400, detail="Excel export not available (openpyxl not installed)")

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "MAUDE Events"

        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for row_idx, event in enumerate(events, 2):
            ws.cell(row=row_idx, column=1, value=event["mdr_report_key"])
            ws.cell(row=row_idx, column=2, value=event["report_number"])
            ws.cell(row=row_idx, column=3, value=event["date_received"])
            ws.cell(row=row_idx, column=4, value=event["date_of_event"])
            ws.cell(row=row_idx, column=5, value=event["event_type"])
            ws.cell(row=row_idx, column=6, value=event["manufacturer"])
            ws.cell(row=row_idx, column=7, value=event["product_code"])

        # Adjust column widths
        column_widths = [15, 20, 15, 15, 12, 50, 12]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=maude_events_{date.today().isoformat()}.xlsx"
            },
        )

    # Default: CSV format
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for event in events:
        writer.writerow([
            event["mdr_report_key"],
            event["report_number"],
            event["date_received"],
            event["date_of_event"],
            event["event_type"],
            event["manufacturer"],
            event["product_code"],
        ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=maude_events_{date.today().isoformat()}.csv"
        },
    )


@router.get("/{mdr_report_key}", response_model=EventDetail)
async def get_event(mdr_report_key: str):
    """Get detailed information for a single event."""
    query_service = QueryService()
    event = query_service.get_event_detail(mdr_report_key)

    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    return event
